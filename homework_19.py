
import csv
import openpyxl

wb = openpyxl.Workbook()
wb.create_sheet(title='Первый лист', index=0)
sheet = wb['Первый лист']

row_index = 1
ind = 1
with open('hw_18.csv', encoding='utf-8') as file:
    file_reader = csv.reader(file)

    for row in file_reader:
        x = row.pop(2)

        person = 'Person' + str(ind-1)
        if person == 'Person0':
            person = ''
        cell = sheet.cell(1, ind)
        cell.value = person
        ind = ind + 1

        for col_index, value in enumerate(row):
            column = col_index + 1
            cell = sheet.cell(column+1, row_index)
            cell.value = value.lower()
        row_index = row_index + 1

wb.save('hw_19.xlsx')
