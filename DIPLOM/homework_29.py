

import openpyxl
import openpyxl as xl
from datetime import *
from dateutil.relativedelta import *


def skl(w, lst):
    f1 = lambda a: (a % 100) // 10 != 1 and a % 10 == 1
    f2 = lambda a: (a % 100) // 10 != 1 and a % 10 in [2, 3, 4]
    return lst[0] if f1(w) else lst[1] if f2(w) else lst[2]


# поиск и отображение информации
def search_inf():
    wb = xl.load_workbook('hw_29.xlsx', read_only=True)
    sheet = wb.active
    search_name = input('Введите запрос: ')
    search_name = str(search_name.upper())
    row_max = sheet.max_row
    print('Всего записей в базе: ', row_max)

    # формирование строки

    for row in sheet.rows:
        person = []
        for cell in row:
            item = str(cell.value)
            person.append(str(item))

        # поиск по строке и формирование списка для печати

        for x in person:
            if search_name in x:

                # подсчет возраста

                date_birthday = datetime.strptime(person[3], '%d.%m.%Y')
                if person[4] == 'None':
                    last_date = datetime.now()
                else:
                    last_date = datetime.strptime(person[4], '%d.%m.%Y')
                y = relativedelta(last_date, date_birthday).years

                # склонение год, года, лет

                years = ['год', 'года', 'лет']
                yres = f'{y} {skl(y, years)}' if y else ''
                person_1 = person[1]
                if person[0] == 'None':
                    person_2 = ''
                else:
                    person_2 = person[0]
                if person[2] == 'None':
                    person_3 = ''
                else:
                    person_3 = person[2]
                if person[5] == 'М':
                    person_4 = ' мужчина.'
                else:
                    person_4 = ' женщина.'
                if person[5] == 'М':
                    person_5 = ' Родился: '
                else:
                    person_5 = ' Родилась: '
                person_6 = person[3]
                if person[4] == 'None':
                    person_7 = ''
                else:
                    person_7 = person[4]
                    if person[5] == 'М':
                        person_7 = ' Умер: ' + person_7 + '.'
                    else:
                        person_7 = ' Умерла: ' + person_7 + '.'
                print('- ', person_1, ' ', person_2, ' ', person_3, ' ', yres, ',',\
                      person_4, person_5, person_6, '.', person_7, sep='')


# ввод и верификация
def vhod() :
    wb = openpyxl.load_workbook('hw_29.xlsx')
    sheet = wb.active
    while True:
        vvod = input('Введите код операции: \t 1 - ввести новую запись;  \t 2 - поиск по  записям; \
         \t 0 - окончание работ;\n \t ?')
        if vvod == '1':
            while True:
                name = input('Введите фамилию')
                name = name.upper()
                name_2 = input('Введите имя')
                name_2 = name_2.upper()
                if name_2 == '':
                    print('Не введено имя, повторите ввод!!!')
                    continue
                name_3 = input('Введите отчество')
                name_3 = name_3.upper()
                while True:
                    name_4 = input('Введите дату рождения: ЧЧ.ММ.ГГГГ : ')
                    name_4 = name_4.replace('/', '.')
                    name_4 = name_4.replace(' ', '.')
                    name_4 = name_4.replace('-', '.')
                    try:
                        valid_date = datetime.strptime(name_4, '%d.%m.%Y')
                    except ValueError:
                        print('Не введена корректно дата рождения, повторите ввод!!!')
                        continue
                    break
                while True:
                    name_5 = input('Введите дату смерти')
                    name_5 = name_5.replace('/', '.')
                    name_5 = name_5.replace(' ', '.')
                    name_5 = name_5.replace('-', '.')
                    if name_5 == '':
                        break
                    else:
                        try:
                            valid_date = datetime.strptime(name_5, '%d.%m.%Y')
                        except ValueError:
                            print('Не введена корректно дата смерти, повторите ввод!!!')
                            continue
                        break
                name_6 = input('Введите пол: м/ж')
                name_6 = name_6.upper()
                if name_6 != 'М' and name_6 != 'Ж':
                    print('Не введен пол (м/ж), повторите ввод!!!')
                    continue
                sheet.append([name, name_2, name_3, name_4, name_5, name_6])
                person = list([name, name_2, name_3, name_4, name_5, name_6])
                print('Введена информация: ', person)
                wb.save('hw_29.xlsx')
                break
        elif vvod == '2':
            search_inf()
        elif vvod == '0':
            break
        if vvod.isdigit() and int(vvod) <= 0 and int(vvod) >= 2:
            print('Bam нужно вводить число от 0 до 2')


vhod()
